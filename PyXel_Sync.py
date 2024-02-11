    # -*- coding: utf-8 -*-
"""
/***************************************************************************
Name			 	 : PyXel Sync
Description          : Use QRMS
Date                 : 11/Feb/24
copyright            : (C) 2024 by J.D.M.P.E
email                : 
 ***************************************************************************/

/***************************************************************************
 *                                                                         *
 *   This program is free software; you can redistribute it and/or modify  *
 *   it under the terms of the GNU General Public License as published by  *
 *   the Free Software Foundation; either version 2 of the License, or     *
 *   (at your option) any later version.                                   *
 *                                                                         *
 ***************************************************************************/
 This script initializes the plugin, making it known to QGIS.
 """

from qgis.gui import QgsMapLayerComboBox,QgsMapToolEmitPoint,QgsRubberBand
from qgis.core import QgsMapLayerProxyModel,QgsMapLayer,QgsProject,QgsPointXY
from qgis.core import QgsProject, QgsVectorLayer, QgsField, QgsFeature, QgsGeometry, QgsVectorFileWriter, QgsWkbTypes,QgsRectangle,QgsProcessingFeatureSourceDefinition
from qgis.PyQt.QtWidgets import QDockWidget, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QCheckBox, QWidget, QPushButton,QDialog,QComboBox
from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtGui import (QIcon, QColor, QFont)
from qgis.PyQt.QtCore import (QSettings, QTranslator, qVersion, QCoreApplication, Qt, QVariant, QUrl,QDir)
from qgis.PyQt.QtWidgets import QMessageBox,QLineEdit,QLabel,QAction
from qgis.core import QgsCoordinateReferenceSystem, QgsProject
from qgis.gui import QgsProjectionSelectionDialog
import csv,os,traceback
from osgeo import ogr, gdal, osr
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSignal
import processing
from openpyxl import Workbook, load_workbook
from os.path import isfile
import traceback
import os
from qgis.core import *
from qgis.gui import *
from qgis.gui import QgsLayerPropertiesWidget,QgsSingleSymbolRendererWidget, QgsRendererWidget, QgsAttributeTableView, QgsAttributeTableModel, QgsAttributeTableFilterModel
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import (QMainWindow,QMenu,QPushButton, QMessageBox, QDialog,QSlider,QColorDialog,QFileDialog, QTableWidget, QVBoxLayout, QTableWidgetItem)
from openpyxl import Workbook, load_workbook
from os.path import isfile
import traceback
import os
from qgis.core import *
from qgis.gui import *
from qgis.gui import QgsLayerPropertiesWidget,QgsSingleSymbolRendererWidget, QgsRendererWidget, QgsAttributeTableView, QgsAttributeTableModel, QgsAttributeTableFilterModel
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import (QMainWindow,QMenu,QPushButton, QMessageBox, QDialog,QSlider,QColorDialog,QFileDialog, QTableWidget, QVBoxLayout, QTableWidgetItem)
OPEN_DIALOGS = []
current_dir = os.path.dirname(os.path.abspath(__file__))

class pyxelSync:
    """QGIS Plugin Starting"""
    
    def __init__(self, iface):
        
        #Define the qgis plugin env-------------------------------------------------------------------------------
        self.iface = iface
        self.plugin_dir = os.path.dirname(__file__)
        # initialize locale
        locale = QSettings().value('locale/userLocale')[0:2]
        locale_path = os.path.join(
            self.plugin_dir,
            'i18n',
            'PyXelSync_{}.qm'.format(locale))
            
        if os.path.exists(locale_path):
            self.translator = QTranslator()
            self.translator.load(locale_path)

            if qVersion() > '4.3.3':
                QCoreApplication.installTranslator(self.translator)

        # Declare instance attributes
        self.actions = []
        self.menu = self.tr(u'&PyXel_Sync')
        # TODO: We are going to let the user set this up in a future iteration
        self.toolbar = self.iface.addToolBar(u'PyXel_Sync')
        self.toolbar.setObjectName(u'PyXel_Sync')
        icon_path= os.path.join(current_dir, 'PyXel_Sync.ico')
        #End of the define the qgis plugin env-------------------------------------------------------------------------------


    def tr(self, message):
        """Get the translation for a string using Qt translation API.

        We implement this ourselves since we do not inherit QObject.

        :param message: String for translation.
        :type message: str, QString

        :returns: Translated version of message.
        :rtype: QString
        """
        # noinspection PyTypeChecker,PyArgumentList,PyCallByClass
        return QCoreApplication.translate('PyXel_Sync', message)

    def add_action(
            self,
            icon_path,
            text,
            callback,
            enabled_flag=True,
            add_to_menu=True,
            add_to_toolbar=True,
            status_tip=None,
            whats_this=None,
            parent=None):

        # Create the dialog (after translation) and keep reference

        icon = QIcon(icon_path)
        action = QAction(icon, text, parent)
        action.triggered.connect(callback)
        action.setEnabled(enabled_flag)

        if status_tip is not None:
            action.setStatusTip(status_tip)

        if whats_this is not None:
            action.setWhatsThis(whats_this)

        if add_to_toolbar:
            self.toolbar.addAction(action)

        if add_to_menu:
            self.iface.addPluginToMenu(
                self.menu,
                action)

        self.actions.append(action)

        return action
        
    def unload(self):
        """Removes the plugin menu item and icon from QGIS GUI."""
        for action in self.actions:
            self.iface.removePluginMenu(
                self.tr(u'&PyXel_Sync'),
                action)
            self.iface.removeToolBarIcon(action)
            # remove the toolbar
        del self.toolbar

    def initGui(self):
        """Create the menu entries and toolbar icons inside the QGIS GUI."""
        icon_path=os.path.normpath(os.path.join(os.path.dirname(__file__), 'PyXel_Sync.gif'))
        self.clickApi=self.add_action(
            icon_path,
            text=self.tr(u'PyXel_Sync'),
            callback=self.run,
            parent=self.iface.mainWindow())     
        self.dlg = selectLayer(self.iface)

        
    def run(self):
        self.dlg.show()

class AttributeDialog(QDialog):
    def __init__(self, canvas, layer):
        super(AttributeDialog, self).__init__()
        self.canvas= canvas
        self.layer = layer
        self.isEditable = False
        self.setWindowTitle('속성편집')
        self.resize(1200, 700)
        # Ensure this instance is not garbage collected
        OPEN_DIALOGS.append(self)

        # Add FID field if it does not exist
        self.add_fid_field()

        self.layout = QVBoxLayout(self)
        
        self.table = QTableWidget(self)
        # Set header color to gray
        header_style = "QHeaderView::section { background-color: gray; }"
        self.table.horizontalHeader().setStyleSheet(header_style)
        self.table.verticalHeader().setStyleSheet(header_style)
        
        self.layout.addWidget(self.table)
        
        self.editButton = QPushButton("Edit Attributes", self)
        self.editButton.clicked.connect(self.toggle_edit)
        
        self.saveButton = QPushButton("Save", self)
        self.saveButton.clicked.connect(self.save_attributes)
        self.saveButton.setEnabled(False)  # Initially disabled
        
        self.layout.addWidget(self.editButton)
        self.layout.addWidget(self.saveButton)
        
        # Add Export to Excel and Import from Excel buttons
        self.exportExcelButton = QPushButton("Export to Excel", self)
        self.exportExcelButton.clicked.connect(self.export_to_excel)
        self.layout.addWidget(self.exportExcelButton)
        
        # Import from Excel button setup
        self.importExcelButton = QPushButton("Import from Excel", self)
        self.importExcelButton.clicked.connect(self.import_from_excel)
        self.layout.addWidget(self.importExcelButton)
        
        self.setLayout(self.layout)

        # Connect the row change signal to our custom method
        self.table.currentCellChanged.connect(self.on_row_change)
        
        self.populate_table()

    def add_fid_field(self):
        fid_field_name = 'fid'
        if fid_field_name not in [field.name() for field in self.layer.fields()]:
            self.layer.startEditing()
            self.layer.dataProvider().addAttributes([QgsField(fid_field_name, QVariant.Int)])
            self.layer.commitChanges()
            
            # Populate FID field with consecutive numbers
            self.layer.startEditing()
            for idx, feature in enumerate(self.layer.getFeatures()):
                feature.setAttribute(fid_field_name, idx)
                self.layer.updateFeature(feature)
            self.layer.commitChanges()


    def delete_rows(self,table,start_row, end_row):
        # Delete rows from the end to the start
        for row in range(end_row, start_row - 1, -1):
            table.removeRow(row)

    def import_from_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel files (*.xlsx)")
        if path:
            
            workbook = load_workbook(filename=path)
            sheet = workbook.active
    
            self.table.clear()  # Clear existing data and headers
  
            # Set headers based on Excel file's first row
            headers = [sheet.cell(row=1, column=col+1).value or '' for col in range(sheet.max_column)]
            self.table.setHorizontalHeaderLabels(headers)
    
            # Load data into the table from the second row of Excel file
            for row in range(2, sheet.max_row + 1):
                table_row = row - 2  # Adjust for zero-based indexing and header row
                print('table_row:',table_row)
                self.table.insertRow(table_row)
                for col in range(1, sheet.max_column + 1):
                    print('col:', col)
                    cell_value = sheet.cell(row=row, column=col).value
                    cell_value_str = '' if cell_value is None else str(cell_value)
                    self.table.setItem(table_row, col - 1, QTableWidgetItem(cell_value_str))
            self.allRowcount = self.table.rowCount()
            self.delete_rows(self.table,int(self.allRowcount/2),self.allRowcount)
  
            # Enable the Save button
            #self.editButton.setText("Stop Editing" if self.isEditable else "Edit Attributes")
            self.saveButton.setEnabled(True)
            #self.saveButton.setEnabled(self.isEditable)
            
            QMessageBox.information(self, "Import Successful", "Data imported from Excel successfully.")

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel files (*.xlsx)")
        if path:
            workbook = Workbook()
            sheet = workbook.active
    
            # Add column headers
            for column in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(column)
                if header is not None:  # Check if header item exists
                    sheet.cell(row=1, column=column+1, value=header.text())
                else:
                    sheet.cell(row=1, column=column+1, value='Header{}'.format(column))
    
            # Add table data
            for row in range(self.table.rowCount()):
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    if item is not None:  # Check if the item is not None
                        sheet.cell(row=row+2, column=column+1, value=item.text())
                    else:
                        sheet.cell(row=row+2, column=column+1, value='')
    
            workbook.save(path)
            QMessageBox.information(self, "Export Successful", "Data exported to Excel successfully.")

    def on_row_change(self, currentRow, currentColumn, previousRow, previousColumn):
        # Assuming the FID is in the first column
        fid_item = self.table.item(currentRow, 0)
        if fid_item:
            fid = int(fid_item.text())
            self.zoom_to_feature(fid)

    def zoom_to_feature(self, fid):
        feature = self.layer.getFeature(fid)
        if feature.isValid():  # Check if the feature with the given FID exists
            # Zoom to feature
            #self.parent.project.mapCanvas().setExtent(feature.geometry().boundingBox())
            self.canvas.setExtent(feature.geometry().boundingBox())
            self.canvas.refresh()

    def toggle_edit(self):
        self.isEditable = not self.isEditable
    
        # Iterate through items and set them as editable or not
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is None:
                    item = QTableWidgetItem()  # Create a new item if it doesn't exist
                    self.table.setItem(row, col, item)
    
                if self.isEditable:
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                else:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        
        self.editButton.setText("Stop Editing" if self.isEditable else "Edit Attributes")
        self.saveButton.setEnabled(self.isEditable)


    def populate_table(self):
        if not isinstance(self.layer, QgsVectorLayer):
            return
    
        fields = self.layer.fields()
        
        # Add 'fid' column as the first column
        self.table.insertColumn(0)
        self.table.setHorizontalHeaderItem(0, QTableWidgetItem('fid'))
    
        # Populate other columns, excluding 'fid'
        field_names = [field.name() for field in fields if field.name() != 'fid']
        for i, field_name in enumerate(field_names):
            self.table.insertColumn(i + 1)
            self.table.setHorizontalHeaderItem(i + 1, QTableWidgetItem(field_name))
    
        # Populate rows with feature data
        for feature in self.layer.getFeatures():
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
    
            # Set 'fid' value
            self.table.setItem(row_position, 0, QTableWidgetItem(str(feature.id())))
    
            # Set other field values
            for i, field_name in enumerate(field_names):
                value = feature[field_name]
                if value is not None:
                    if str(value)is "NULL":
                        item = QTableWidgetItem(str(''))
                    else:
                        item = QTableWidgetItem(str(value))
                else:
                    item = QTableWidgetItem(str(''))
                #item = QTableWidgetItem(str(value) if value is not None else '')
                self.table.setItem(row_position, i + 1, item)
    
        # Remove duplicate 'fid' column if exists
        fid_col_count = self.table.columnCount()
        for i in range(fid_col_count):
            if self.table.horizontalHeaderItem(i).text() == 'fid' and i != 0:
                self.table.removeColumn(i)
                break

    def save_attributes(self):
        success = self.layer.startEditing()
    
        if not success:
            QMessageBox.warning(self, "Error", "Unable to start editing the layer.")
            return
    
        for row in range(self.table.rowCount()):
            feature_id = int(self.table.item(row, 0).text())
            feature = self.layer.getFeature(feature_id)
    
            if feature.isValid():
                for col in range(1, self.table.columnCount()):
                    field_name = self.table.horizontalHeaderItem(col).text()
                    new_value = self.table.item(row, col).text()
    
                    # Check the field data type
                    field_index = feature.fields().indexOf(field_name)
                    field_type = feature.fields().field(field_index).type()
    
                    try:
                        if new_value == '':
                            if field_type == QVariant.Int or field_type == QVariant.Double:
                                new_value = None  # Use None for NULL values in numeric fields
                            else:
                                new_value = new_value  # Keep the empty string for other data types
                        elif field_type == QVariant.Int:
                            new_value = int(new_value) if new_value else None
                        elif field_type == QVariant.Double:
                            new_value = float(new_value) if new_value else None
                        # Add other data type conversions as needed
    
                        feature.setAttribute(field_name, new_value)
    
                    except ValueError:
                        QMessageBox.warning(self, "Data Type Error", f"Wrong data type for attribute {col} of feature {row}: {new_value}")
                        self.layer.rollBack()
                        return

                self.layer.updateFeature(feature)
    
        if not self.layer.commitChanges():
            QMessageBox.warning(self, "Error", "Failed to commit changes.")
        else:
            QMessageBox.information(self, "Success", "Attributes updated successfully.")
        
        #delete fid field for joins
        fid_field_name = 'fid'
        if fid_field_name in [field.name() for field in self.layer.fields()]:
            self.layer.startEditing()
            fid_index = self.layer.fields().indexFromName(fid_field_name)
            self.layer.dataProvider().deleteAttributes([fid_index])
            self.layer.commitChanges()

    def closeEvent(self, event):
        # Remove FID field when closing the dialog
        fid_field_name = 'fid'
        if fid_field_name in [field.name() for field in self.layer.fields()]:
            self.layer.startEditing()
            fid_index = self.layer.fields().indexFromName(fid_field_name)
            self.layer.dataProvider().deleteAttributes([fid_index])
            self.layer.commitChanges()
        super().closeEvent(event)

class selectLayer(QDialog):
    def __init__(self, iface):
        super().__init__()

        self.iface = iface
        self.setWindowTitle("레이어 속성편집")  # Set the window title
        self.resize(200, 100)
        
        self.layout = QVBoxLayout(self)

        # Layer selection combo box using QgsMapLayerComboBox
        self.layer_combo = QgsMapLayerComboBox()
        self.layer_combo.setFilters(QgsMapLayerProxyModel.PointLayer | QgsMapLayerProxyModel.LineLayer | QgsMapLayerProxyModel.PolygonLayer)
        self.populate_layer_combo()
        self.layout.addWidget(QLabel("Select Layer:"))
        self.layout.addWidget(self.layer_combo)

        # Execute and Cancel buttons
        self.execute_button = QPushButton("Execute")
        self.cancel_button = QPushButton("Cancel")
        self.layout.addWidget(self.execute_button)
        self.layout.addWidget(self.cancel_button)

        self.execute_button.clicked.connect(self.on_execute)
        self.cancel_button.clicked.connect(self.close)

    def populate_layer_combo(self):
        # Add all layers from the map canvas to the combo box
        layers = QgsProject.instance().mapLayers().values()
        for layer in layers:
            if layer.type() == QgsMapLayer.VectorLayer:  # Only add vector layers
                self.layer_combo.addItem(layer.name(), layer)

        # Optionally, you can set the current layer as the active layer in the combo box
        current_layer = self.iface.activeLayer()
        if current_layer and current_layer.type() == QgsMapLayer.VectorLayer:
            self.layer_combo.setLayer(current_layer)

    def closeEvent(self, event):
        self.is_running = False
        super().closeEvent(event)

    def show_attribute_window(self, layer):
        dlg = AttributeDialog(self.iface.mapCanvas(), layer)
        #dlg.exec_()
        dlg.show()

    def on_execute(self):

        try:
            # Usage example:
            layer = self.layer_combo.currentLayer()
            if not layer.isValid():
                print("Layer failed to load!")
                return
            else :
                self.show_attribute_window(layer) # Show the attribute window for the first selected feature
                    
        except:
            print(traceback.print_exc())

    def closeEvent(self, event):
        self.is_running = False
        #if self.layer_combo.currentLayer():
        #    self.layer_combo.currentLayer().commitChanges()  # Save changes
        
        # Optionally, if you want to commit changes immediately
        #self.update_layer.commitChanges()

        super().closeEvent(event)

